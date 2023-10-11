import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common import by
from selenium.webdriver.common.keys import Keys

# Ścieżka do pliku Excel z nazwami uczelni
from webdriver_manager.chrome import ChromeDriverManager

data_link = 'C:\Python Projects\Scrapper\src\InputExcell\InputAlioth.xlsx'
excel_file_path = data_link

# Otwórz plik Excel
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb.active

# Inicjalizacja przeglądarki Chrome
# driver = webdriver.Chrome(executable_path='sciezka_do_chromedriver.exe') # Podaj ścieżkę do chromedriver.exe
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.get("https://www.google.com")

# Przechodzimy przez każdy wiersz pliku Excel
for row in sheet.iter_rows(min_row=2, values_only=True):
    nazwa_uczelni = row[0]  # Zakładamy, że nazwa uczelni znajduje się w pierwszej kolumnie

    # Przeszukaj Google w poszukiwaniu strony internetowej uczelni
    driver.get("https://www.google.com")
    search_box = driver.find_element(by.NAME, value="q")
    search_box.send_keys(nazwa_uczelni + " strona internetowa")
    search_box.send_keys(Keys.RETURN)

    # Sprawdzamy, czy znaleziono wyniki
    results = driver.find_elements_by_css_selector("div.g")
    website_found = False
    website_url = "BRAK"

    for result in results:
        link = result.find_element_by_css_selector("a")
        description = result.find_element_by_css_selector("div.IsZvec")
        if nazwa_uczelni.lower() in description.text.lower():
            website_url = link.get_attribute("href")
            website_found = True
            break

    # Zapisz wynik do pliku Excel
    sheet.cell(row=sheet.cell(row=row[0]).row, column=sheet.cell(row=row[0]).column + 1, value=website_url)

# Zapisz zmiany w pliku Excel
wb.save(data_link)

# Zakończ przeglądarkę
driver.quit()